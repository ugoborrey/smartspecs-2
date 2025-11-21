#!/usr/bin/env python3
"""
PIM attribute enrichment via OpenAI Responses API (gpt-5-nano, batch + structured outputs).

Commands:
  prepare --input <xlsx> --output <batch.jsonl>
  send --batch-file <batch.jsonl>
  merge --batch-id <batch_id> --input <xlsx> --out-dir <dir>

Prereqs:
  - OPENAI_API_KEY in env (loadable via dotenv if present)
  - pip install openpyxl python-dotenv openai
"""

import argparse
import json
import os
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
import requests

try:
    from dotenv import load_dotenv
except ImportError:  # optional dependency
    def load_dotenv():
        return None

MODEL = "gpt-5-nano"


# --- Excel parsing ---------------------------------------------------------
def load_workbook(path: Path):
    return openpyxl.load_workbook(path, data_only=True)


def load_lov(wb: openpyxl.Workbook) -> Dict[Tuple[str, str], List[str]]:
    """(pim_class, attr_code) -> allowedValues list."""
    sheet = wb["AttributesPossibleValues"]
    lov: Dict[Tuple[str, str], List[str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pim_class = row[0]
        attr_code = row[1]
        if not pim_class or not attr_code:
            continue
        values = [str(v).strip() for v in row[2:] if v not in (None, "")]
        lov[(str(pim_class).strip(), str(attr_code).strip())] = values
    return lov


def build_header_maps(sheet: openpyxl.worksheet.worksheet.Worksheet):
    headers = list(sheet.iter_rows(min_row=1, max_row=1))[0]
    regex = re.compile(r"^(\d+)_([\s\S]+)$")
    attr_cols: List[Tuple[int, str, str, str]] = []  # (idx, col_letter, code, name)
    info_cols: List[Tuple[int, str]] = []  # (idx, header)
    for idx, cell in enumerate(headers):
        val = cell.value if cell.value is not None else ""
        m = regex.match(str(val))
        if m:
            attr_cols.append((idx, cell.column_letter, m.group(1), m.group(2)))
        else:
            info_cols.append((idx, str(val)))
    return attr_cols, info_cols


def color_code(cell) -> str:
    if not cell or not cell.fill or not cell.fill.fgColor:
        return ""
    fg = cell.fill.fgColor
    if fg.rgb:
        return fg.rgb.upper()
    if fg.indexed is not None:
        return f"indexed:{fg.indexed}"
    return ""


def is_applicable(cell) -> bool:
    c = color_code(cell)
    if c in {"FF000000", "000000"}:
        return False
    return True  # grey, white, or no fill


def normalize_key(val: str) -> str:
    return (val or "").strip()


def fetch_pkh_data(brand: str, market_ref: str) -> Dict[str, Any]:
    """
    Fetch canonical payload from PKH API. Returns dict with either
    {"product": {...}} or {"not_found": True, ...}
    """
    load_dotenv()
    base_url = os.getenv("PKH_BASE_URL", "http://127.0.0.1:8000")
    url = f"{base_url}/products_canonical"
    params = {
        "brand": (brand or "").strip(),
        "manufacturer_reference": (market_ref or "").strip(),
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        if resp.status_code == 404:
            return {"not_found": True, "brand": brand, "market_ref": market_ref}
        resp.raise_for_status()
        data = resp.json()
        # keep only the product block
        return {"product": data.get("product"), "brand": brand, "market_ref": market_ref}
    except Exception as exc:
        return {"not_found": True, "brand": brand, "market_ref": market_ref, "error": str(exc)}


def extract_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    export = wb["Export"]
    lov = load_lov(wb)
    attr_cols, info_cols = build_header_maps(export)

    # Helpful lookups for brand/market_ref/pim class
    header_index = {h: idx for idx, h in info_cols}
    brand_idx = header_index.get("Brand Name(s)")
    market_idx = header_index.get("Market ref")
    pim_class_idx = header_index.get("PIM M Class")

    rows_out: List[Dict[str, Any]] = []
    for excel_row_index, row in enumerate(export.iter_rows(min_row=2), start=2):
        pim_class = str(row[pim_class_idx].value).strip() if pim_class_idx is not None and row[pim_class_idx].value else ""
        if not pim_class:
            continue

        brand = ""
        if brand_idx is not None and row[brand_idx].value is not None:
            brand = normalize_key(str(row[brand_idx].value))

        market_ref = ""
        if market_idx is not None and row[market_idx].value is not None:
            market_ref = normalize_key(str(row[market_idx].value))

        attributes = OrderedDict()
        for idx, col_letter, code, name in attr_cols:
            cell = row[idx]
            if not is_applicable(cell):
                continue
            key = (pim_class, code)
            if key in lov:
                attributes[name] = {"allowedValues": lov[key]}
            else:
                attributes[name] = None

        if not attributes:
            continue

        rows_out.append(
            {
                "row_index": excel_row_index,
                "pim_class": pim_class,
                "brand": brand,
                "market_ref": market_ref,
                "attributes": attributes,
            }
        )
    return rows_out


# --- LLM payload helpers ---------------------------------------------------
def build_response_format(attributes: OrderedDict) -> Dict[str, Any]:
    """
    Crée un schéma strict où chaque attribut avec LoV est contraint par enum.
    """
    attr_properties: Dict[str, Any] = {}
    required_attrs: List[str] = []
    for name, info in attributes.items():
        required_attrs.append(name)
        if isinstance(info, dict) and "allowedValues" in info:
            allowed = list(info["allowedValues"])
            enum_values = allowed + [None]
            attr_properties[name] = {"type": ["string", "null"], "enum": enum_values}
        else:
            attr_properties[name] = {"type": ["string", "null"]}

    return {
        "type": "json_schema",
        "name": "pim_enrichment",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "attributes": {
                    "type": "object",
                    "properties": attr_properties,
                    "required": required_attrs,
                    "additionalProperties": False,
                }
            },
            "required": ["attributes"],
            "additionalProperties": False,
        },
    }


def build_system_prompt() -> str:
    return (
        "Tu enrichis les attributs PIM uniquement à partir des données du Product Knowledge Hub fournies. "
        "Si allowedValues est présent, retourne strictement l'une de ces valeurs. "
        "Si aucune valeur n'est supportée par les données PKH, renvoie null. "
        "N'invente pas de valeur en dehors des allowedValues."
    )


def make_custom_id(sheet_row_index: int) -> str:
    return f"pim-row-{sheet_row_index}"


# --- Batch steps -----------------------------------------------------------
def step_prepare(input_file: Path, batch_file: Path) -> None:
    # Resolve output path relative to this script if not absolute
    if not batch_file.is_absolute():
        batch_file = Path(__file__).parent / batch_file

    rows = extract_rows(input_file)
    system_prompt = build_system_prompt()
    requests: List[str] = []
    not_found: List[Dict[str, Any]] = []

    for row in rows:
        pkh_payload = fetch_pkh_data(row["brand"], row["market_ref"])
        if pkh_payload.get("not_found"):
            not_found.append(
                {
                    "row_index": row["row_index"],
                    "brand": row["brand"],
                    "market_ref": row["market_ref"],
                    "reason": pkh_payload.get("error", "not_found"),
                }
            )
            continue  # skip this request
        json_schema = build_response_format(row["attributes"])
        user_payload = {
            "data_source_pkh": pkh_payload,
            "pim_attributes_to_enrich": row["attributes"],
        }
        req = {
            "custom_id": make_custom_id(row["row_index"]),
            "method": "POST",
            "url": "/v1/responses",
            "body": {
                "model": MODEL,
                "input": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                ],
                "text": {"format": json_schema},
                "max_output_tokens": 16000,
            },
        }
        requests.append(json.dumps(req, ensure_ascii=False))

    batch_file.parent.mkdir(parents=True, exist_ok=True)
    with batch_file.open("w", encoding="utf-8") as f:
        for r in requests:
            f.write(r + "\n")
    print(f"Écrit {len(requests)} requêtes dans {batch_file}")

    report_path = batch_file.parent / "not_found_report.json"
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(not_found, f, ensure_ascii=False, indent=2)
    print(f"Non trouvés PKH : {len(not_found)} (rapport : {report_path})")


def step_send(batch_file: Path) -> None:
    from openai import OpenAI

    load_dotenv()
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    bf = client.files.create(file=open(batch_file, "rb"), purpose="batch")
    job = client.batches.create(
        input_file_id=bf.id,
        endpoint="/v1/responses",
        completion_window="24h",
    )
    print(f"Batch lancé : {job.id}")


def parse_response_line(line: str) -> Dict[str, Any]:
    obj = json.loads(line)
    cid = obj.get("custom_id")
    body = obj.get("response", {}).get("body", {})
    out = body.get("output") or []
    payload = None
    for item in out:
        if item.get("type") != "message":
            continue
        for c in item.get("content", []):
            if c.get("type") == "output_json":
                payload = c.get("json")
                break
            if c.get("type") == "output_text":
                try:
                    payload = json.loads(c.get("text", "") or "{}")
                except json.JSONDecodeError:
                    payload = None
        if payload:
            break
    return {"custom_id": cid, "payload": payload}


def step_merge(batch_id: str, input_file: Path, out_dir: Path) -> None:
    from openai import OpenAI

    load_dotenv()
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    job = client.batches.retrieve(batch_id)
    if job.status != "completed":
        raise SystemExit(f"Batch {batch_id} pas terminé, statut={job.status}")
    content = client.files.content(job.output_file_id).text

    result_map: Dict[str, Any] = {}
    for line in content.splitlines():
        if not line.strip():
            continue
        parsed = parse_response_line(line)
        if parsed.get("custom_id") and parsed.get("payload"):
            result_map[parsed["custom_id"]] = parsed["payload"]

    out_dir.mkdir(parents=True, exist_ok=True)
    # Option 1: garder les JSON (diagnostic rapide)
    written = 0
    for cid, payload in result_map.items():
        out_path = out_dir / f"{cid}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        written += 1
    if written:
        print(f"Fichiers JSON écrits: {written} dans {out_dir}")
    else:
        print("Aucun payload à écrire")

    # Option 2: réinjecter dans un Excel enrichi
    enriched_path = out_dir / "pim_enriched.xlsx"
    if result_map:
        wb = openpyxl.load_workbook(input_file)
        export = wb["Export"]
        attr_cols, _ = build_header_maps(export)
        # map attr name -> column index
        name_to_idx = {name: idx for idx, _, _, name in attr_cols}

        for excel_row_index, row in enumerate(export.iter_rows(min_row=2), start=2):
            cid = make_custom_id(excel_row_index)
            payload = result_map.get(cid)
            if not payload:
                continue
            attrs = payload.get("attributes", {})
            for attr_name, val in attrs.items():
                if val is None:
                    continue
                col_idx = name_to_idx.get(attr_name)
                if col_idx is None:
                    continue
                row[col_idx].value = val
        wb.save(enriched_path)
        print(f"Excel enrichi écrit : {enriched_path}")
    else:
        print("Aucun enrichissement à appliquer dans Excel")


# --- CLI -------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="PIM enrichment batch (gpt-5-nano, Responses API)")
    sub = parser.add_subparsers(dest="cmd")

    p1 = sub.add_parser("prepare")
    p1.add_argument("--input", required=True, type=Path, help="Fichier source XLSX PIM")
    p1.add_argument("--output", required=True, type=Path, help="Fichier batch .jsonl")

    p2 = sub.add_parser("send")
    p2.add_argument("--batch-file", required=True, type=Path, help="Fichier batch .jsonl")

    p3 = sub.add_parser("merge")
    p3.add_argument("--batch-id", required=True, help="ID du batch job")
    p3.add_argument("--input", required=True, type=Path, help="Fichier source XLSX PIM")
    p3.add_argument("--out-dir", required=True, type=Path, help="Dossier de sortie JSON")

    args = parser.parse_args()
    if args.cmd == "prepare":
        step_prepare(args.input, args.output)
    elif args.cmd == "send":
        step_send(args.batch_file)
    elif args.cmd == "merge":
        step_merge(args.batch_id, args.input, args.out_dir)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
